import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from datetime import datetime
import threading
import subprocess
import sys
import json
import os

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "testing" / "test-case.xlsx"
SELENIUM_SCRIPT = ROOT / "scripts" / "selenium_autotest.py"
SUMMARY_PATH = ROOT / "testing" / "selenium-summary.json"


def normalize(text):
    return " ".join(str(text or "").strip().lower().split())


class TestRunnerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Attendance Test Runner")
        # Fit within the screen while keeping title bar/taskbar visible.
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        target_w, target_h = 1900, 1000
        margin_w, margin_h = 40, 80
        fit_w = min(target_w, max(800, screen_w - margin_w))
        fit_h = min(target_h, max(600, screen_h - margin_h))
        # Place the window near the top-left so the title bar is fully visible.
        self.geometry(f"{fit_w}x{fit_h}+0+0")
        self.resizable(False, False)
        # Outer border for the whole GUI window
        self.configure(highlightthickness=2, highlightbackground="#000000", bg="#f8fafc")
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        title = ttk.Label(self, text="Test Runner (Excel)", font=("Segoe UI", 14, "bold"))
        title.grid(row=0, column=0, columnspan=2, sticky="w", **pad)

        preview_outer = tk.Frame(self, bg="#9ca3af", padx=12, pady=12)
        preview_outer.grid(row=0, column=2, rowspan=12, sticky="n", padx=(10, 28), pady=10)
        preview_inner = tk.Frame(preview_outer, bg="#cbd5e1", padx=6, pady=6)
        preview_inner.grid(row=0, column=0, sticky="n")
        preview_frame = ttk.LabelFrame(preview_inner, text="Live Browser Preview", labelanchor="nw")
        try:
            preview_frame.configure(labelwidget=ttk.Label(preview_frame, text="Live Browser Preview", font=("Segoe UI", 10, "bold")))
        except Exception:
            pass
        preview_frame.grid(row=0, column=0, sticky="n")
        preview_frame.configure(padding=12, borderwidth=2, relief="solid")
        self.preview_w = 650
        self.preview_h = 650
        self.preview_canvas = tk.Canvas(
            preview_frame,
            width=self.preview_w,
            height=self.preview_h,
            bg="#f8fafc",
            highlightthickness=0,
            bd=0,
            relief="flat",
        )
        self.preview_canvas.grid(row=0, column=0, padx=8, pady=8)
        self._preview_img = None
        self._preview_running = False
        self._preview_mtime = None
        self._set_preview_black()

        ttk.Label(self, text="S.NO (optional)").grid(row=1, column=0, sticky="w", **pad)
        self.sno_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.sno_var, width=30).grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(self, text="Requirement description").grid(row=2, column=0, sticky="w", **pad)
        self.req_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.req_var, width=60).grid(row=2, column=1, sticky="w", **pad)

        ttk.Label(self, text="Test inputs").grid(row=3, column=0, sticky="w", **pad)
        self.inputs_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.inputs_var, width=60).grid(row=3, column=1, sticky="w", **pad)

        ttk.Label(self, text="Expected result").grid(row=4, column=0, sticky="w", **pad)
        self.expected_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.expected_var, width=60).grid(row=4, column=1, sticky="w", **pad)

        self.add_if_missing = tk.BooleanVar(value=True)
        ttk.Checkbutton(self, text="Add new row if not found", variable=self.add_if_missing).grid(
            row=5, column=1, sticky="w", **pad
        )

        self.show_browser = tk.BooleanVar(value=False)

        btn_frame = ttk.Frame(self)
        btn_frame.grid(row=7, column=0, columnspan=2, sticky="w", **pad)
        ttk.Button(btn_frame, text="Run / Save", command=self.run_and_save).grid(row=0, column=0, padx=4)
        ttk.Button(btn_frame, text="Run Selenium", command=self.run_selenium).grid(row=0, column=1, padx=4)
        ttk.Button(btn_frame, text="Clear", command=self.clear_fields).grid(row=0, column=2, padx=4)

        results_frame = ttk.LabelFrame(self, text="Results (after Run Selenium)")
        results_frame.grid(row=8, column=0, columnspan=2, sticky="we", padx=10, pady=6)

        ttk.Label(results_frame, text="Actual result").grid(row=0, column=0, sticky="w", padx=10, pady=6)
        self.result_actual_var = tk.StringVar(value="")
        ttk.Entry(results_frame, textvariable=self.result_actual_var, width=60).grid(row=0, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(results_frame, text="Status").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        self.status_var = tk.StringVar(value="--")
        ttk.Entry(results_frame, textvariable=self.status_var, width=30, state="readonly").grid(row=1, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(results_frame, text="Observation").grid(row=2, column=0, sticky="w", padx=10, pady=6)
        self.obs_var = tk.StringVar(value="")
        ttk.Entry(results_frame, textvariable=self.obs_var, width=60, state="readonly").grid(row=2, column=1, sticky="w", padx=10, pady=6)

        log_frame = ttk.LabelFrame(self, text="Automation Log (shows pages opened)")
        log_frame.grid(row=9, column=0, columnspan=2, sticky="we", padx=10, pady=6)
        self.log_text = tk.Text(log_frame, height=7, width=86, wrap="word")
        self.log_text.grid(row=0, column=0, sticky="we", padx=10, pady=6)
        self.log_text.configure(state="disabled")

        self.status_label = ttk.Label(self, text="Ready.", foreground="#2563eb")
        self.status_label.grid(row=10, column=0, columnspan=2, sticky="w", **pad)

        note = ttk.Label(
            self,
            text="Tip: Close Excel before running to avoid file lock.",
            foreground="#6b7280",
        )
        note.grid(row=11, column=0, columnspan=2, sticky="w", **pad)

    def clear_fields(self):
        self.sno_var.set("")
        self.req_var.set("")
        self.inputs_var.set("")
        self.expected_var.set("")
        self.result_actual_var.set("")
        self.status_var.set("--")
        self.obs_var.set("")
        self.status_label.config(text="Cleared.", foreground="#2563eb")
        self._append_log("Cleared form.")
        self._set_preview_black()

    def run_and_save(self):
        req = self.req_var.get().strip()
        inputs = self.inputs_var.get().strip()
        expected = self.expected_var.get().strip()
        actual = self.result_actual_var.get().strip()
        sno = self.sno_var.get().strip()

        if not req:
            messagebox.showerror("Required", "Requirement description cannot be blank.")
            return
        if not expected:
            messagebox.showerror("Required", "Expected result cannot be blank.")
            return
        if not actual:
            messagebox.showerror("Required", "Actual result cannot be blank. Run Selenium or type it in Results.")
            return

        try:
            wb = load_workbook(XLSX_PATH)
            ws = wb.active
        except Exception as exc:
            messagebox.showerror("Excel Error", f"Could not open Excel file:\n{exc}")
            return

        # Map headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if isinstance(val, str):
                headers[val.strip().lower()] = col

        col_sno = headers.get("s.no") or headers.get("s.no.") or headers.get("sno")
        col_req = headers.get("requirement description")
        col_inputs = headers.get("test inputs")
        col_expected = headers.get("expected result")
        col_actual = headers.get("actual result")
        col_status = headers.get("status(pass/fail)")
        col_obs = headers.get("observation")

        missing = [k for k, v in [
            ("Requirement description", col_req),
            ("Test inputs", col_inputs),
            ("Expected result", col_expected),
            ("Actual result", col_actual),
            ("Status(Pass/Fail)", col_status),
            ("Observation", col_obs),
        ] if v is None]
        if missing:
            messagebox.showerror("Excel Error", f"Missing columns in Excel: {', '.join(missing)}")
            return

        # Find row by S.NO or requirement description
        target_row = None
        if sno and col_sno:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_sno).value
                if val is None:
                    continue
                if str(val).strip() == sno:
                    target_row = row
                    break

        if target_row is None:
            req_norm = normalize(req)
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_req).value
                if normalize(val) == req_norm:
                    target_row = row
                    break

        if target_row is None:
            if not self.add_if_missing.get():
                messagebox.showerror("Not Found", "No matching row found for S.NO or requirement.")
                return
            target_row = ws.max_row + 1
            if col_sno and sno:
                ws.cell(row=target_row, column=col_sno, value=sno)

        ws.cell(row=target_row, column=col_req, value=req)
        ws.cell(row=target_row, column=col_inputs, value=inputs)
        ws.cell(row=target_row, column=col_expected, value=expected)
        ws.cell(row=target_row, column=col_actual, value=actual)

        status = "Pass" if normalize(expected) == normalize(actual) else "Fail"
        ws.cell(row=target_row, column=col_status, value=status)

        if status == "Fail":
            obs = "Expected and actual results did not match after run."
        else:
            obs = ""
        ws.cell(row=target_row, column=col_obs, value=obs)

        try:
            wb.save(XLSX_PATH)
        except Exception as exc:
            messagebox.showerror("Save Error", f"Could not save Excel file:\n{exc}")
            return

        ts = datetime.now().strftime("%H:%M:%S")
        self.status_label.config(text=f"Saved at {ts}. Status: {status}", foreground="#16a34a" if status == "Pass" else "#dc2626")
        self._set_result(actual, status, obs)
        self._append_log(f"Saved results to Excel. Status: {status}.")

    def run_selenium(self):
        if not SELENIUM_SCRIPT.exists():
            messagebox.showerror("Missing Script", f"Not found: {SELENIUM_SCRIPT}")
            return

        self.status_label.config(text="Running Selenium... Please wait.", foreground="#2563eb")
        self._append_log("Starting Selenium automation...")
        self._append_log(f"Opening app at http://localhost:3000")
        try:
            self.update_idletasks()
            canvas_x = self.preview_canvas.winfo_rootx()
            canvas_y = self.preview_canvas.winfo_rooty()
            canvas_w = self.preview_canvas.winfo_width()
            canvas_h = self.preview_canvas.winfo_height()
            self._append_log(f"Preview canvas at x={canvas_x}, y={canvas_y}, w={canvas_w}, h={canvas_h}")
        except Exception:
            pass
        self._set_preview_black()

        def _run():
            try:
                cmd = [sys.executable, str(SELENIUM_SCRIPT)]
                env = dict(**dict(os.environ))
                # Disable live screenshot preview; keep the panel empty/black.
                env["PREVIEW_PATH"] = ""
                env["PREVIEW_SIZE"] = ""
                env["HEADLESS"] = "0"
                # Treat 1 cm as 1 px for all adjustments.
                cm = lambda v: int(v)
                # Fit the Edge window inside the preview canvas, accounting for window chrome.
                chrome_w = 150
                chrome_h = 200
                # 50cm ≈ 1900px
                # Increase width by 75cm on the right.
                fit_w = cm(500 + 380)
                # Increase height by 5cm ≈ 190px.
                # Reduce height by 10cm.
                # Reduce height by 1cm from current.
                fit_h = max(200, self.preview_h - chrome_h + cm(395))
                env["BROWSER_SIZE"] = f"{fit_w},{fit_h}"
                # Position Edge over the preview panel so it looks embedded.
                self.update_idletasks()
                try:
                    # Align Edge to the black canvas area for a tighter fit.
                    self.update_idletasks()
                    canvas_x = self.preview_canvas.winfo_rootx()
                    canvas_y = self.preview_canvas.winfo_rooty()
                    # Align Edge window with the preview canvas area.
                    # Move Edge down by ~0.5cm.
                    # Move Edge right by 150px.
                    env["BROWSER_POS"] = f"{canvas_x + cm(180) + cm(2) + cm(3) - cm(11) - cm(8)},{canvas_y + cm(20) + cm(2) + cm(1)}"
                except Exception:
                    pass
                self._append_log("Live preview enabled (Edge window visible).")
                result = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True, env=env)
                if result.returncode == 0:
                    self.status_label.config(text="Selenium run completed. Excel and summary updated.", foreground="#16a34a")
                    if result.stdout:
                        self._append_log(result.stdout.strip())
                    if result.stderr:
                        self._append_log(result.stderr.strip())
                    self._load_latest_result()
                    self._stop_preview_poll()
                    self._set_preview_black()
                else:
                    err = (result.stderr or result.stdout or "").strip()
                    short = err.splitlines()[-1] if err else "Unknown error"
                    self.status_label.config(text=f"Selenium failed: {short}", foreground="#dc2626")
                    if err:
                        self._append_log(err)
                    self._stop_preview_poll()
                    self._set_preview_black()
                    messagebox.showerror("Selenium Failed", err or "Unknown error")
            except Exception as exc:
                self.status_label.config(text=f"Selenium failed: {exc}", foreground="#dc2626")
                self._stop_preview_poll()
                self._set_preview_black()
                messagebox.showerror("Selenium Failed", str(exc))

        threading.Thread(target=_run, daemon=True).start()

    def _load_latest_result(self):
        sno = self.sno_var.get().strip()
        if not sno:
            return
        if not SUMMARY_PATH.exists():
            return
        try:
            data = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
        except Exception:
            return
        entry = data.get(str(int(sno))) if str(sno).isdigit() else data.get(sno)
        if not entry:
            return
        actual = entry.get("actual", "")
        status = entry.get("status", "")
        obs = entry.get("observation", "")
        self._set_result(actual, status or "--", obs)

    def _set_result(self, actual, status, obs):
        self.result_actual_var.set(actual)
        self.status_var.set(status or "--")
        self.obs_var.set(obs or "")

    def _append_log(self, message):
        if not message:
            return
        self.log_text.configure(state="normal")
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _draw_preview_border(self):
        # Keep the border fully inside the canvas so the right edge doesn't get clipped.
        pad = 2
        self.preview_canvas.create_rectangle(
            pad, pad, self.preview_w - pad, self.preview_h - pad,
            outline="#64748b", width=2
        )

    def _preview_path(self):
        return ROOT / "testing" / "selenium-preview.png"

    def _set_preview_black(self):
        self.preview_canvas.delete("all")
        self.preview_canvas.create_rectangle(0, 0, self.preview_w, self.preview_h, fill="#f8fafc", outline="#f8fafc")
        self._draw_preview_border()

    def _load_preview_image(self):
        path = self._preview_path()
        if not path.exists():
            return
        mtime = path.stat().st_mtime
        if self._preview_mtime == mtime:
            return
        try:
            img = tk.PhotoImage(file=str(path))
        except Exception:
            return
        # Fit image inside 900x900 box using repeated subsample if needed.
        max_w, max_h = self.preview_w, self.preview_h
        while img.width() > max_w or img.height() > max_h:
            img = img.subsample(2, 2)
        self._preview_img = img
        self._preview_mtime = mtime
        self.preview_canvas.delete("all")
        x = (self.preview_w - img.width()) // 2
        y = (self.preview_h - img.height()) // 2
        self.preview_canvas.create_image(x, y, image=self._preview_img, anchor="nw")
        self._draw_preview_border()

    def _poll_preview(self):
        if not self._preview_running:
            return
        self._load_preview_image()
        self.after(600, self._poll_preview)

    def _start_preview_poll(self):
        if self._preview_running:
            return
        self._preview_running = True
        self._poll_preview()

    def _stop_preview_poll(self):
        self._preview_running = False


if __name__ == "__main__":
    app = TestRunnerApp()
    app.mainloop()
