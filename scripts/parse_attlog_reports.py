import calendar
import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd

UPLOADS_DIR = Path(r"C:\Users\DELL\Desktop\attendance-system\uploads")
FILENAME_RE = re.compile(
    r"^(January|February|March|April|May|June|July|August|September|October|November|December)_(\d{4})\.(xls|xlsx)$"
)
MONTH_TO_NUM = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
}
SHEET_NAME = "Att.log report"
TIME_RE = re.compile(r"\d{2}:\d{2}")


def discover_report_files():
    if not UPLOADS_DIR.exists():
        return []

    items = []
    for path in UPLOADS_DIR.iterdir():
        if not path.is_file():
            continue
        m = FILENAME_RE.match(path.name)
        if not m:
            continue
        year = int(m.group(2))
        month = MONTH_TO_NUM[m.group(1)]
        items.append((year, month, path))
    items.sort()
    return [p for _, _, p in items]


def to_minutes(t):
    h, m = map(int, t.split(":"))
    return h * 60 + m


def to_hhmm(mins):
    mins = max(0, mins)
    return f"{mins // 60:02d}:{mins % 60:02d}"


def normalize_punch_times(times, target_count=4, similarity_window_mins=3):
    if len(times) <= 1:
        return times

    minute_times = [to_minutes(t) for t in times]

    # First pass: collapse obvious duplicate/near-duplicate adjacent punches.
    compact = [minute_times[0]]
    for m in minute_times[1:]:
        if m - compact[-1] <= similarity_window_mins:
            continue
        compact.append(m)

    # If still over target (e.g., 5 punches), remove one from the closest pair
    # until only the required number remain.
    while len(compact) > target_count:
        closest_idx = 0
        closest_gap = compact[1] - compact[0]
        for i in range(1, len(compact) - 1):
            gap = compact[i + 1] - compact[i]
            if gap < closest_gap:
                closest_gap = gap
                closest_idx = i

        # For last pair, keep the later punch (likely final logout).
        if closest_idx == len(compact) - 2:
            remove_idx = closest_idx
        else:
            # In start/middle pairs, keep the first and drop the second.
            remove_idx = closest_idx + 1
        compact.pop(remove_idx)

    return [to_hhmm(m) for m in compact]


def parse_punch_cell(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    times = TIME_RE.findall(text)
    if not times:
        return None
    times = normalize_punch_times(times)

    in_time = times[0]
    logout_for_lunch = "--"
    login_from_lunch = "--"

    if len(times) >= 4:
        logout_for_lunch = times[1]
        login_from_lunch = times[2]
        logout = times[-1]
    elif len(times) == 3:
        logout_for_lunch = times[1]
        logout = times[2]
    elif len(times) == 2:
        logout = times[1]
    else:
        logout = times[0]

    work_mins = to_minutes(logout) - to_minutes(in_time)
    if logout_for_lunch != "--" and login_from_lunch != "--":
        work_mins -= max(0, to_minutes(login_from_lunch) - to_minutes(logout_for_lunch))

    return {
        "inTime": in_time,
        "logoutForLunch": logout_for_lunch,
        "loginFromLunch": login_from_lunch,
        "logout": logout,
        "outTime": logout,
        "work": to_hhmm(work_mins),
        "idle": "00:00",
    }


def read_sheet_rows(path):
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[SHEET_NAME]
        rows = [tuple(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
    else:
        wb = xlrd.open_workbook(str(path))
        sh = wb.sheet_by_name(SHEET_NAME)
        rows = [tuple(sh.row_values(i)) for i in range(sh.nrows)]
    return rows


def parse_month_key(rows):
    header = rows[2] if len(rows) > 2 else ()
    text = " ".join(str(v) for v in header if v not in (None, ""))
    m = re.search(r"(\d{4})-(\d{2})-\d{2}\s*~", text)
    if not m:
        raise ValueError(f"Unable to parse month range from header: {text}")
    return f"{m.group(1)}-{m.group(2)}"


def build_default_month(year, month):
    days = calendar.monthrange(year, month)[1]
    records = []
    for d in range(1, days + 1):
        dow = datetime(year, month, d).weekday()  # Mon=0
        status = "weekend" if dow >= 5 else "absent"
        records.append({"day": d, "status": status})
    return records


def parse_file(path):
    rows = read_sheet_rows(path)
    month_key = parse_month_key(rows)
    year, month = map(int, month_key.split("-"))

    day_row = rows[3]
    day_columns = {}
    for idx, val in enumerate(day_row):
        if isinstance(val, (int, float)):
            d = int(val)
            if 1 <= d <= 31:
                day_columns[idx] = d

    monthly = {}
    current_emp = None

    for row in rows[4:]:
        first = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        if first == "ID:":
            raw_id = row[2] if len(row) > 2 else None
            if raw_id in (None, ""):
                current_emp = None
                continue
            emp_id = str(raw_id).strip()
            current_emp = emp_id
            if current_emp not in monthly:
                monthly[current_emp] = build_default_month(year, month)
            continue

        if not current_emp:
            continue

        row_has_punch = any(parse_punch_cell(row[c]) for c in day_columns.keys() if c < len(row))
        if not row_has_punch:
            continue

        for c, day in day_columns.items():
            if c >= len(row):
                continue
            parsed = parse_punch_cell(row[c])
            if not parsed:
                continue
            rec = monthly[current_emp][day - 1]
            rec.update(parsed)
            rec["status"] = "present"

    return month_key, monthly


def main():
    report_files = discover_report_files()
    if not report_files:
        raise FileNotFoundError(
            f"No valid report files found in {UPLOADS_DIR}. "
            "Expected names like October_2025.xls or November_2025.xlsx"
        )

    merged = {}
    for fp in report_files:
        month_key, data = parse_file(fp)
        for emp_id, records in data.items():
            merged.setdefault(emp_id, {})[month_key] = records

    out_path = Path("src/data/attendance_reports.json")
    out_path.write_text(json.dumps(merged, indent=2), encoding="utf-8")
    print(f"Wrote {out_path} with {len(merged)} employees")


if __name__ == "__main__":
    main()
